"""xlsx 读取 — 每个 sheet 一段截断 CSV + 维度. data_only=True 拿缓存的公式值.

为什么 CSV 不是 Markdown table: 实际表格往往 30+ 列 100+ 行, MD table 渲染巨长
且 LLM 抓不准列对齐. CSV 单行紧凑, agent 自行解析方便.
"""

from __future__ import annotations

import csv
import io
from pathlib import Path

from pentaloom.capabilities.file._models import FileReadResult, FileWarning


def _sheet_to_csv(ws, max_rows: int, max_cols: int) -> tuple[str, dict[str, int]]:
    """把单 sheet 截断 max_rows × max_cols 渲染成 CSV, 返回 (text, stats)."""
    total_rows = ws.max_row or 0
    total_cols = ws.max_column or 0
    take_rows = min(total_rows, max_rows)
    take_cols = min(total_cols, max_cols)

    buf = io.StringIO()
    writer = csv.writer(buf)
    # ws.iter_rows max_row 是包含, 列同理
    rows_iter = ws.iter_rows(
        min_row=1, max_row=take_rows, min_col=1, max_col=take_cols, values_only=True
    )
    for row in rows_iter:
        # None → 空串, 其它 str() 化. 防 csv writer 对 datetime 报 type 错.
        writer.writerow(["" if v is None else str(v) for v in row])

    return buf.getvalue().rstrip("\n"), {
        "total_rows": total_rows,
        "total_cols": total_cols,
        "shown_rows": take_rows,
        "shown_cols": take_cols,
    }


def read_xlsx(
    path: Path,
    *,
    sheet: str | None = None,
    max_rows: int = 200,
    max_cols: int = 30,
) -> FileReadResult:
    from openpyxl import load_workbook  # type: ignore[import-not-found]

    # data_only=True: 公式被替换为 Excel 上次打开时的计算结果. 没打开过的文件
    # 会拿到 None — agent 应该感知到 (但 openpyxl 不区分 "未求值" 和 "真为空"),
    # 不专门 warning, 留给 file_verify 后续扩展.
    wb = load_workbook(str(path), data_only=True, read_only=True)
    try:
        sheet_names = list(wb.sheetnames)
        targets: list[str]
        if sheet is not None:
            if sheet not in sheet_names:
                raise ValueError(
                    f"sheet={sheet!r} 不存在, 可用 sheet: {sheet_names}"
                )
            targets = [sheet]
        else:
            targets = sheet_names

        sections: list[str] = []
        warnings: list[FileWarning] = []
        for name in targets:
            ws = wb[name]
            body, stats = _sheet_to_csv(ws, max_rows=max_rows, max_cols=max_cols)
            header = (
                f"## Sheet: {name}  "
                f"(shown {stats['shown_rows']}×{stats['shown_cols']} of "
                f"{stats['total_rows']}×{stats['total_cols']})"
            )
            sections.append(header + ("\n" + body if body else "\n_(empty)_"))

            hidden_rows = stats["total_rows"] - stats["shown_rows"]
            hidden_cols = stats["total_cols"] - stats["shown_cols"]
            if hidden_rows > 0 or hidden_cols > 0:
                warnings.append(FileWarning(
                    kind="truncated",
                    location=f"sheet={name}",
                    message=(
                        f"truncated: {hidden_rows} rows / {hidden_cols} cols hidden "
                        f"(扩 max_rows / max_cols 或换 sheet 再读)"
                    ),
                ))

        text = "\n\n".join(sections).strip()
        meta: dict[str, int | str | list[str]] = {
            "sheets": len(sheet_names),
            "sheet_names": sheet_names,
            "chars": len(text),
        }
        return FileReadResult(
            path=str(path),
            kind="xlsx",
            text=text,
            warnings=warnings,
            meta=meta,
        )
    finally:
        wb.close()
