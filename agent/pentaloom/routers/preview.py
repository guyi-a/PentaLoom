"""文件预览 endpoint — 给前端右栏 FilePreviewPanel 用.

3 个 endpoint, 都按 session 校 path (sandbox ∪ mounted_dirs):
  - GET /fs/preview/stat?session_id=&path=     → metadata (含 is_binary_guess)
  - GET /fs/preview/text?session_id=&path=&max_bytes=     → 截断文本
  - GET /fs/preview/file?session_id=&path=     → 二进制流 (img/iframe/video src 用)

设计:
  - 预留 token 参数 — 现在直接 sid+path query 鉴权, 加短 TTL token 时只改这层
  - HITL 不审 — 用户点击触发的 UI action, 不是 LLM 工具
  - text 限制 max_bytes (默认 512KB), 二进制 (NUL 字节探测) 拒绝走 text endpoint
  - file endpoint 注入 Content-Type + Content-Disposition: inline (PDF iframe 必需,
    否则浏览器触发下载)
  - symlink resolve 后 relative_to(allowed_root) 已在 path_scope helper 内做
"""

from __future__ import annotations

from pathlib import Path

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import FileResponse
from pydantic import BaseModel

from pentaloom.infra.path_scope import resolve_session_scoped_path

router = APIRouter(prefix="/fs/preview", tags=["fs"])

# 文本读默认上限 512KB. 超了截断 + UI 提示, 跟 krow 同款.
DEFAULT_MAX_BYTES = 512 * 1024
# 上限保护 — 防 caller 传特别大的 max_bytes 卡内存.
HARD_MAX_BYTES = 4 * 1024 * 1024

# binary 探测读前 4KB 看 NUL 字节. 跟很多终端 / git diff 的策略一致.
BINARY_PROBE_BYTES = 4 * 1024


# ext → mime. 浏览器原生渲 (img/iframe/video/audio) 用; office 二进制 (docx/pptx)
# 也走这里, 前端 fetch as ArrayBuffer 给 docx-preview / pptx-renderer 库渲.
# 不在 map 内的 ext 走 stat + text endpoint, 不走 file endpoint.
MIME_MAP: dict[str, str] = {
    # image
    "png": "image/png",
    "jpg": "image/jpeg",
    "jpeg": "image/jpeg",
    "gif": "image/gif",
    "webp": "image/webp",
    "svg": "image/svg+xml",
    "bmp": "image/bmp",
    "ico": "image/x-icon",
    # pdf
    "pdf": "application/pdf",
    # video
    "mp4": "video/mp4",
    "webm": "video/webm",
    "mov": "video/quicktime",
    "m4v": "video/mp4",
    "mkv": "video/x-matroska",
    # audio
    "mp3": "audio/mpeg",
    "wav": "audio/wav",
    "ogg": "audio/ogg",
    "m4a": "audio/mp4",
    "aac": "audio/aac",
    "flac": "audio/flac",
    "opus": "audio/opus",
    # office binary — 浏览器不直渲, 前端 fetch as ArrayBuffer 喂给 docx-preview /
    # pptx-renderer 库做客户端渲染. xlsx 不在这里 (走 /fs/preview/office 出结构化).
    "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "pptx": "application/vnd.openxmlformats-officedocument.presentationml.presentation",
}


class PreviewMeta(BaseModel):
    path: str            # 规范化后绝对路径
    name: str
    size: int
    ext: str             # 不含点 . 的小写后缀; 无后缀返 ""
    mtime: float         # unix ts
    is_directory: bool
    is_binary_guess: bool  # 读前 4KB 探测 NUL 字节; True → 不应走 text endpoint


class TextPreviewResult(BaseModel):
    content: str
    truncated: bool      # 文件大于 max_bytes 被截断
    size: int            # 文件原始字节数


def _ext_of(p: Path) -> str:
    """basename 最后一个 .  之后的小写 ext, 无后缀返 ''. .env / Dockerfile 这种无 ext 文件
    返 '' (前端 classifyFile 用 basename 兜底)."""
    suf = p.suffix.lstrip(".").lower()
    return suf


def _is_binary_buffer(buf: bytes) -> bool:
    """看前 N 字节有没有 NUL — 文本文件几乎不会有 NUL, 二进制有概率有.

    不区分 utf-16 之类含 NUL 的"宽字节文本" — 这种文件本来就不该走 text preview.
    """
    return b"\x00" in buf


@router.get("/stat", response_model=PreviewMeta)
async def preview_stat(
    session_id: str = Query(..., description="当前 session id"),
    path: str = Query(..., description="绝对路径"),
) -> PreviewMeta:
    target = await resolve_session_scoped_path(session_id, path)
    stat = target.stat()

    is_dir = target.is_dir()
    is_binary = False
    if not is_dir:
        try:
            with target.open("rb") as f:
                head = f.read(BINARY_PROBE_BYTES)
            is_binary = _is_binary_buffer(head)
        except OSError:
            # 读不出 binary 探测就当 binary, 防前端真二进制硬走 text
            is_binary = True

    return PreviewMeta(
        path=str(target),
        name=target.name,
        size=stat.st_size,
        ext=_ext_of(target),
        mtime=stat.st_mtime,
        is_directory=is_dir,
        is_binary_guess=is_binary,
    )


@router.get("/text", response_model=TextPreviewResult)
async def preview_text(
    session_id: str = Query(...),
    path: str = Query(...),
    max_bytes: int = Query(DEFAULT_MAX_BYTES, ge=1, le=HARD_MAX_BYTES),
) -> TextPreviewResult:
    target = await resolve_session_scoped_path(
        session_id, path, require_file=True,
    )

    stat = target.stat()
    # 真二进制不能走 text — 防大二进制 utf-8 decode + replace 卡前端
    with target.open("rb") as f:
        head = f.read(BINARY_PROBE_BYTES)
    if _is_binary_buffer(head):
        raise HTTPException(400, "binary file; use /fs/preview/file instead")

    truncated = stat.st_size > max_bytes
    if truncated:
        # 已读 head, 接着读到 max_bytes 即可
        with target.open("rb") as f:
            buf = f.read(max_bytes)
    else:
        with target.open("rb") as f:
            buf = f.read()

    return TextPreviewResult(
        content=buf.decode("utf-8", errors="replace"),
        truncated=truncated,
        size=stat.st_size,
    )


# xlsx 解析上限. 跟 krow 同款 — 大表整个 cell 解析 cost 高, cap 防卡.
XLSX_MAX_SHEETS = 20
XLSX_MAX_ROWS = 200
XLSX_MAX_COLS = 50
XLSX_MAX_CELLS = 50_000
XLSX_SIZE_LIMIT = 20 * 1024 * 1024


class XlsxCellStyle(BaseModel):
    bold: bool | None = None
    italic: bool | None = None
    font_size: float | None = None
    color: str | None = None       # CSS color, e.g. "#3d5a80"
    bg_color: str | None = None
    align: str | None = None       # "left" | "center" | "right"
    valign: str | None = None      # "top" | "middle" | "bottom"


class XlsxCell(BaseModel):
    text: str
    style: XlsxCellStyle | None = None


class XlsxMerge(BaseModel):
    start_row: int
    start_col: int
    end_row: int
    end_col: int


class XlsxSheet(BaseModel):
    name: str
    row_count: int       # 真实总行数
    col_count: int       # 真实总列数
    rows: list[list[XlsxCell]]
    merges: list[XlsxMerge]
    truncated: bool      # 是否被 cap 截断


class XlsxWorkbookPreview(BaseModel):
    sheets: list[XlsxSheet]
    active_sheet_index: int
    truncated: bool      # 整体是否被 sheet 数 / cell 总数 cap 截断
    size: int            # 文件字节数


def _xlsx_color(color: object) -> str | None:
    """openpyxl Color → "#rgb". rgb 格式可能是 "FFRRGGBB" (含 alpha) 或 "RRGGBB"."""
    if color is None:
        return None
    rgb = getattr(color, "rgb", None)
    if not rgb or not isinstance(rgb, str):
        return None
    # AARRGGBB → RRGGBB
    if len(rgb) == 8:
        rgb = rgb[2:]
    if len(rgb) != 6:
        return None
    # 全黑 / 全白 不返 — 跟默认色重复, 渲染只会增噪
    if rgb.upper() in ("000000", "FFFFFF"):
        return None
    return "#" + rgb.lower()


def _xlsx_cell_style(cell) -> XlsxCellStyle | None:
    """openpyxl Cell → XlsxCellStyle. 只抽 5 项视觉 (bold/italic/size/color/bg+align).
    复杂样式 (border / number_format / wrap_text) 第一版不抽."""
    style = XlsxCellStyle()
    has_any = False

    font = cell.font
    if font:
        if font.bold:
            style.bold = True
            has_any = True
        if font.italic:
            style.italic = True
            has_any = True
        if font.size and font.size != 11:  # 11 是 Excel 默认
            style.font_size = float(font.size)
            has_any = True
        font_color = _xlsx_color(font.color)
        if font_color:
            style.color = font_color
            has_any = True

    fill = cell.fill
    if fill and fill.patternType == "solid":
        bg = _xlsx_color(fill.fgColor) or _xlsx_color(fill.bgColor)
        if bg:
            style.bg_color = bg
            has_any = True

    align = cell.alignment
    if align:
        if align.horizontal in ("center", "right", "left"):
            style.align = align.horizontal
            has_any = True
        if align.vertical in ("top", "center", "bottom"):
            style.valign = "middle" if align.vertical == "center" else align.vertical
            has_any = True

    return style if has_any else None


@router.get("/office/xlsx", response_model=XlsxWorkbookPreview)
async def preview_xlsx(
    session_id: str = Query(...),
    path: str = Query(...),
) -> XlsxWorkbookPreview:
    """xlsx 结构化预览 — 用 openpyxl 抽每个 cell 的 text + 样式 + merges, 前端 HTML
    table 真渲 (sticky header / 字体颜色 / 背景 / 合并单元格).

    cap: 20 sheets × 200 rows × 50 cols, 整 workbook 50k cells. 超了 truncated=True.
    """
    target = await resolve_session_scoped_path(session_id, path, require_file=True)
    ext = _ext_of(target)
    if ext not in ("xlsx", "xlsm"):
        raise HTTPException(415, f"xlsx endpoint 不支持 ext={ext!r}")
    stat = target.stat()
    if stat.st_size > XLSX_SIZE_LIMIT:
        raise HTTPException(413, f"file size {stat.st_size} exceeds xlsx limit {XLSX_SIZE_LIMIT}")

    # 延迟 import — openpyxl 启动慢
    from openpyxl import load_workbook  # type: ignore[import-not-found]

    # data_only=True: 公式被替换为 Excel 上次打开时的计算结果 (没打开过文件会拿 None).
    # read_only=False 因为 read_only 模式拿不到 cell.font / cell.fill 等样式信息.
    wb = load_workbook(str(target), data_only=True, read_only=False)
    try:
        all_sheet_names = list(wb.sheetnames)
        sheet_names = all_sheet_names[:XLSX_MAX_SHEETS]
        workbook_truncated = len(all_sheet_names) > XLSX_MAX_SHEETS

        sheets: list[XlsxSheet] = []
        total_cells = 0
        for name in sheet_names:
            ws = wb[name]
            total_rows = ws.max_row or 0
            total_cols = ws.max_column or 0
            take_rows = min(total_rows, XLSX_MAX_ROWS)
            take_cols = min(total_cols, XLSX_MAX_COLS)

            rows: list[list[XlsxCell]] = []
            for r in range(1, take_rows + 1):
                if total_cells >= XLSX_MAX_CELLS:
                    break
                row_cells: list[XlsxCell] = []
                for c in range(1, take_cols + 1):
                    total_cells += 1
                    cell = ws.cell(row=r, column=c)
                    text = "" if cell.value is None else str(cell.value)
                    row_cells.append(XlsxCell(
                        text=text,
                        style=_xlsx_cell_style(cell),
                    ))
                rows.append(row_cells)

            merges: list[XlsxMerge] = []
            for mrange in ws.merged_cells.ranges:
                # openpyxl 的 merge range 用 1-based, 转 0-based 跟前端对齐.
                # 限制在 cap 内才包含.
                start_row = mrange.min_row - 1
                start_col = mrange.min_col - 1
                if start_row >= take_rows or start_col >= take_cols:
                    continue
                merges.append(XlsxMerge(
                    start_row=start_row,
                    start_col=start_col,
                    end_row=min(mrange.max_row - 1, take_rows - 1),
                    end_col=min(mrange.max_col - 1, take_cols - 1),
                ))

            sheets.append(XlsxSheet(
                name=name,
                row_count=total_rows,
                col_count=total_cols,
                rows=rows,
                merges=merges,
                truncated=(
                    total_rows > XLSX_MAX_ROWS
                    or total_cols > XLSX_MAX_COLS
                    or total_cells >= XLSX_MAX_CELLS
                ),
            ))

            if total_cells >= XLSX_MAX_CELLS:
                break

        # 取 active sheet 的 index (落在 sheet_names 内才有意义)
        active_idx = 0
        try:
            active_name = wb.active.title if wb.active else sheet_names[0]
            active_idx = sheet_names.index(active_name) if active_name in sheet_names else 0
        except (AttributeError, ValueError):
            active_idx = 0

        return XlsxWorkbookPreview(
            sheets=sheets,
            active_sheet_index=active_idx,
            truncated=workbook_truncated or total_cells >= XLSX_MAX_CELLS,
            size=stat.st_size,
        )
    finally:
        wb.close()


# zip 解析上限. 不读内容只读目录, 零 zip-bomb 数据风险, 仅靠 entry count cap
# 防"几百万空文件名"这种 metadata bomb.
ZIP_SIZE_LIMIT = 100 * 1024 * 1024
ZIP_MAX_ENTRIES = 5000


class ZipEntry(BaseModel):
    path: str             # zip 内相对路径
    size: int             # 解压后字节数
    compressed_size: int
    is_dir: bool


class ArchivePreview(BaseModel):
    entries: list[ZipEntry]
    total_entries: int    # 真实总数 (可能 > len(entries) 当 truncated)
    truncated: bool
    size: int             # 文件大小


@router.get("/archive/zip", response_model=ArchivePreview)
async def preview_zip(
    session_id: str = Query(...),
    path: str = Query(...),
) -> ArchivePreview:
    """zip 文件结构预览 — 只列 metadata 不解压内容, 零 zip-bomb 数据风险.

    cap: 5000 entries 防 metadata bomb. 文件 100MB. tar/gz 第一版不支持.
    """
    target = await resolve_session_scoped_path(session_id, path, require_file=True)
    ext = _ext_of(target)
    if ext != "zip":
        raise HTTPException(415, f"zip endpoint 不支持 ext={ext!r}")
    stat = target.stat()
    if stat.st_size > ZIP_SIZE_LIMIT:
        raise HTTPException(
            413, f"file size {stat.st_size} exceeds zip limit {ZIP_SIZE_LIMIT}",
        )

    # zipfile 是 stdlib, 不必延迟 import (跟 xlsx 的 openpyxl 不同, 后者很重).
    import zipfile

    entries: list[ZipEntry] = []
    try:
        with zipfile.ZipFile(str(target)) as zf:
            infolist = zf.infolist()
            total = len(infolist)
            for info in infolist[:ZIP_MAX_ENTRIES]:
                # ZipInfo.is_dir() 看 filename 末尾 /. 处理 "目录/" 跟 "文件" 两类.
                entries.append(ZipEntry(
                    path=info.filename,
                    size=info.file_size,
                    compressed_size=info.compress_size,
                    is_dir=info.is_dir(),
                ))
    except zipfile.BadZipFile as e:
        raise HTTPException(422, f"corrupted zip: {e}") from e

    return ArchivePreview(
        entries=entries,
        total_entries=total,
        truncated=total > ZIP_MAX_ENTRIES,
        size=stat.st_size,
    )


# sqlite 解析上限. cap 跟 xlsx 同档 — 数据库单表大概率超 200 行, 截断后给前端
# 提示 "看了前 200, 共 N 行".
SQLITE_SIZE_LIMIT = 50 * 1024 * 1024
SQLITE_MAX_TABLES = 20
SQLITE_MAX_ROWS = 200
SQLITE_MAX_COLS = 50
SQLITE_MAX_CELLS = 50_000


class SqliteTable(BaseModel):
    name: str
    columns: list[str]    # 列名 (不带 type, 第一版只渲名字)
    rows: list[list[str]] # 每行每列的 str 表示, None → ""
    row_count: int        # 真实总行数 (COUNT(*))
    truncated: bool       # 行 > MAX_ROWS 或 列 > MAX_COLS 或 cells 超 budget


class DatabasePreview(BaseModel):
    tables: list[SqliteTable]
    total_tables: int     # 真实总表数 (可能 > len(tables))
    truncated: bool       # tables / cells 超 cap
    size: int


def _quote_sqlite_name(name: str) -> str:
    """SQLite identifier 双引号 quote, 内部 " 转义为 "". 防表名/列名注入."""
    return '"' + name.replace('"', '""') + '"'


@router.get("/office/sqlite", response_model=DatabasePreview)
async def preview_sqlite(
    session_id: str = Query(...),
    path: str = Query(...),
) -> DatabasePreview:
    """sqlite 数据库预览 — 列表名 / 列名 / 前 200 行 / 行数. 只读模式打开防误改.

    cap: 20 tables × 200 rows × 50 cols, 50k cells 整库. 文件 50MB.
    """
    target = await resolve_session_scoped_path(session_id, path, require_file=True)
    ext = _ext_of(target)
    if ext not in ("db", "sqlite", "sqlite3"):
        raise HTTPException(415, f"sqlite endpoint 不支持 ext={ext!r}")
    stat = target.stat()
    if stat.st_size > SQLITE_SIZE_LIMIT:
        raise HTTPException(
            413, f"file size {stat.st_size} exceeds sqlite limit {SQLITE_SIZE_LIMIT}",
        )

    # 延迟 import — 跟 openpyxl 同模式 (sqlite3 是 stdlib 不慢, 但保持一致).
    import sqlite3

    # mode=ro 只读防误改; immutable=1 告诉 sqlite "文件不会被并发改", 跳过 wal /
    # 锁检查, 启动快. 我们读 metadata + 几百行不需要 wal 语义.
    uri = f"file:{target}?mode=ro&immutable=1"
    try:
        conn = sqlite3.connect(uri, uri=True)
    except sqlite3.Error as e:
        raise HTTPException(422, f"corrupted sqlite: {e}") from e

    try:
        # 跳过 sqlite_* 内部表 (sqlite_master / sqlite_sequence 等)
        names_cursor = conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' "
            "AND name NOT LIKE 'sqlite_%' ORDER BY name"
        )
        all_names = [r[0] for r in names_cursor.fetchall()]
        total_tables = len(all_names)
        sample_names = all_names[:SQLITE_MAX_TABLES]

        tables: list[SqliteTable] = []
        total_cells = 0
        for name in sample_names:
            if total_cells >= SQLITE_MAX_CELLS:
                break
            quoted = _quote_sqlite_name(name)

            # 列名 (从 PRAGMA table_info, 字段顺序: cid/name/type/...)
            col_cursor = conn.execute(f"PRAGMA table_info({quoted})")
            all_columns = [r[1] for r in col_cursor.fetchall()]
            columns = all_columns[:SQLITE_MAX_COLS]
            cols_truncated = len(all_columns) > SQLITE_MAX_COLS

            # 总行数
            row_count = conn.execute(f"SELECT COUNT(*) FROM {quoted}").fetchone()[0]

            # 前 200 行 — 只取被 cap 的列
            if columns:
                quoted_cols = ", ".join(_quote_sqlite_name(c) for c in columns)
                rows_cursor = conn.execute(
                    f"SELECT {quoted_cols} FROM {quoted} LIMIT ?",
                    (SQLITE_MAX_ROWS,),
                )
            else:
                # 没列 (空 schema 表) — 跳过 row 取
                rows_cursor = []  # type: ignore[assignment]

            limited_rows: list[list[str]] = []
            for r in rows_cursor:
                if total_cells + len(r) > SQLITE_MAX_CELLS:
                    break
                row_strs = ["" if v is None else str(v) for v in r]
                limited_rows.append(row_strs)
                total_cells += len(row_strs)

            tables.append(SqliteTable(
                name=name,
                columns=columns,
                rows=limited_rows,
                row_count=row_count,
                truncated=(
                    cols_truncated
                    or row_count > SQLITE_MAX_ROWS
                    or total_cells >= SQLITE_MAX_CELLS
                ),
            ))

        return DatabasePreview(
            tables=tables,
            total_tables=total_tables,
            truncated=(
                total_tables > SQLITE_MAX_TABLES
                or total_cells >= SQLITE_MAX_CELLS
            ),
            size=stat.st_size,
        )
    finally:
        conn.close()


@router.get("/file")
async def preview_file(
    session_id: str = Query(...),
    path: str = Query(...),
):
    """返二进制流给 <img|iframe|video|audio src=> 直接 load.

    Content-Type 按 ext mime map 注入. 不在 map 内的 ext 拒 — 防注入意外 mime 给浏览器.
    Content-Disposition: inline 必需 (不然 PDF iframe 触发下载, GPT review 提的).
    """
    target = await resolve_session_scoped_path(
        session_id, path, require_file=True,
    )
    ext = _ext_of(target)
    mime = MIME_MAP.get(ext)
    if mime is None:
        raise HTTPException(
            415, f"file ext {ext!r} not supported by /fs/preview/file (mime allowlist)",
        )
    return FileResponse(
        path=str(target),
        media_type=mime,
        headers={
            "Content-Disposition": "inline",
            "Cache-Control": "no-store",  # 文件可能被改, 别缓存
        },
    )
